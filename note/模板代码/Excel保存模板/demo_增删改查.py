# -*- encoding: utf-8 -*-
# https://www.cnblogs.com/shaosks/p/6098282.html
# ----- https://blog.csdn.net/qq_38486203/article/details/89844146
# ----- https://geek-docs.com/python/python-tutorial/python-openpyxl.html
# ----- https://blog.csdn.net/sinat_28576553/article/details/81275650


# 注意这里只能是xlsx类型的excel
import os
import string
import openpyxl
from xlutils.copy import copy as cop
# 导入字体、边框、颜色以及对齐方式相关库
import xlrd
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

# ------------  读  ----------


# 读取excel---xlsx数据：获取指定sheet数据
#        返回所有数据: (按行输出数据  # 按列输出数据)
def read_xlsx(file_name, Sheet_name):
    """
        读取excel---xlsx数据：指定sheet数据
    :param file_name:  xlsx表名称
    :param Sheet_name: 表内的sheet表名
    :return:
    """
    # 读取excel数据
    file_obj = openpyxl.load_workbook(file_name)

    # 获取指定sheet表
    file_data = file_obj.get_sheet_by_name(Sheet_name)

    # 按行输出数据
    row_data = [[cell.value for cell in row] for row in file_data.rows]

    # 按列输出数据
    columns_data = [[cell.value for cell in column] for column in file_data.columns]

    # 关闭表
    file_obj.close()
    return row_data, columns_data


# 读取整个表数据
# num: 1是按行读取， 2是按列读取
def read_xlsx_all(file_name, num=1):
    """
      读取整个表数据
    :param file_name: 目标文件
    :param num: 1是按行读取， 2是按列读取
    :return:
    """
    # 读取excel数据
    file_obj = openpyxl.load_workbook(file_name)

    # 获取所有sheet子表名 --- list
    sheet_name_list = file_obj.get_sheet_names()

    # 遍历读表
    list_data = []  # 数据list
    for Sheet_name in sheet_name_list:
        print('*'*20, f'当前读取的表：{Sheet_name}', '*'*20)

        # 获取指定sheet表
        file_data = file_obj.get_sheet_by_name(Sheet_name)

        # 判断是按列还是按行读取
        if num == 1:
            # ------------按行
            # 按行输出数据
            row_data = [[cell.value for cell in row] for row in file_data.rows]
            # 数据缓存
            list_data.append(row_data)
        else:
            # ------------按列
            # 按列输出数据
            columns_data = [[cell.value for cell in column] for column in file_data.columns]
            # 数据缓存
            list_data.append(columns_data)
    # 关闭表
    file_obj.close()
    return list_data


# 自己选择指定读取
# num: 1是按行读取， 2是按列读取
def read_xlsx_once(file_name, num=1):
    """
      自己选择指定读取
    :param file_name: 目标文件
    :param num: 1是按行读取， 2是按列读取
    :return:
    """
    # 读取excel数据
    file_obj = openpyxl.load_workbook(file_name)

    # 获取所有sheet子表名 --- list
    sheet_name_list = file_obj.get_sheet_names()

    [print(f'{i}: {sheet_name_list[i]}') for i in range(len(sheet_name_list))]
    print('---------填写相应的序号，确定需要读取的工作表---------')

    # 选着性读取数据
    Sheet_num = []
    try:
        while True:
            num = input('请填写一个哦！：').split()[0]
            Sheet_num.append(int(num))

            a = input('输入stop停止，选择！！否者（enter）继续选择：')
            if a == 'stop':
                break
    except Exception as e:
        print('请正确输入')

    sheet_name_list = [sheet_name_list[i] for i in Sheet_num if i < len(sheet_name_list)]

    # 遍历读表
    list_data = []  # 数据list
    for Sheet_name in sheet_name_list:
        print('*'*20, f'当前读取的表：{Sheet_name}', '*'*20)

        # 获取指定sheet表
        file_data = file_obj.get_sheet_by_name(Sheet_name)

        # 判断是按列还是按行读取
        if num == 1:
            # 按行输出数据
            row_data = [[cell.value for cell in row] for row in file_data.rows]
            # 数据缓存
            list_data.append(row_data)
        else:
            # 按列输出数据
            columns_data = [[cell.value for cell in column] for column in file_data.columns]
            # 数据缓存
            list_data.append(columns_data)
    # 关闭表
    file_obj.close()
    return list_data

# ---------------- 写 ----------------------------


# 创建新表 写入数据 --- 逐行写入
#   一组数据代表 表中一行的数据
def write_xlsx_row(file_path, data, dir_name='数据', sheet_name='Sheet1', num=0):
    """
        创建新表 写入数据 --- 逐行写入
    :param file_path: 保存的工作表名称
    :param data: 需要写入的数据，列表数据
    :param dir_name: 保存数据的目录
    :param sheet_name: 子表名称
    :param num: 0，使用默认工作表   1，自己创建工作表
    :return:
    """
    # 保存数据名
    file_name = os.path.join(dir_name, file_path)

    # 目录是否存在，不存在创建
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)

    # 新建一个Excel文档对象
    wb = openpyxl.Workbook()

    # 工作表
    if num == 0:
        sheet = wb.active
    else:
        sheet = wb.create_sheet(sheet_name)

    # # 设置字体风格为(name)Times New Roman，大小(size)为16，粗体(bold)、斜体(italic)，颜色(color)黑色
    # sheet.font = Font(name='Times New Roman', size=16, bold=True, italic=False, color=colors.BLUE)
    # # 通过参数horizontal和vertical来设置文字在单元格里的对齐方式，此外设置值还可为left和right
    # sheet.alignment = Alignment(horizontal='center', vertical='center')

    # 写数据
    for row in data:
        sheet.append(row)

    # 保存数据
    wb.save(file_name)


# 创建新表 写入数据 --- 逐列写入
#   一组数据代表 表中一列的数据
def write_xlsx_columns(file_path, data, dir_name='数据', sheet_name='Sheet1', num=0):
    """
        创建新表 写入数据 --- 逐列写入
    :param file_path: 保存的工作表名称
    :param data: 需要写入的数据，列表数据
    :param dir_name: 保存数据的目录
    :param sheet_name: 子表名称
    :param num: 0，使用默认工作表   1，自己创建工作表
    :return:
    """
    # 保存数据名
    file_name = os.path.join(dir_name, file_path)

    # 目录是否存在，不存在创建
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)

    # 新建一个Excel文档对象
    wb = openpyxl.Workbook()

    # 工作表
    if num == 0:
        sheet = wb.active
    else:
        sheet = wb.create_sheet(sheet_name)

    # 重新排序
    for i in range(len(data[0])):
        list_column = []
        for j in data:
            list_column.append(j[i])
        sheet.append(list_column)
    # 保存数据
    wb.save(file_name)


# ---------------- 改 --------------

# 修改一行数据 可通过start_row_columns设置从列哪里开始改
def set_excel_row(file_path, Sheet_name, set_rows, data, start_row_columns=1):
    """
        修改某行数据
    :param file_path: 目标excel文件——————-xlsx
    :param Sheet_name: 修改的sheet工作表名
    :param set_rows: 修改行
    :param data:  修改数据
    :param start_row_columns:  从哪列开始修改
    :return:
    """
    # 找到需要更改的表
    work = openpyxl.load_workbook(file_path)

    # 取待修改的工作表
    sheet = work[Sheet_name]

    # 修改目标行数据
    nums = 0
    for i in range(start_row_columns-1, len(data)+start_row_columns-1):
        word = string.ascii_lowercase[i].upper()
        sheet[f'{word}{str(set_rows)}'] = data[nums]
        nums += 1

    work.save(file_path)


# 修改一列数据， 可通过start_columns_row设置从行哪里开始改
def set_excel_columns(file_path, Sheet_name, set_columns, data, start_columns_row=1):
    """
        修改某列数据
    :param file_path: 目标excel文件——————-xlsx
    :param Sheet_name: 修改的sheet工作表名
    :param set_columns: 修改列
    :param start_columns_row: 修改起始行
    :param data:  修改数据
    :return:
    """
    # 找到需要更改的表
    work = openpyxl.load_workbook(file_path)

    # 取待修改的工作表
    sheet = work[Sheet_name]

    # 修改目标行数据
    nums = 0
    for i in range(start_columns_row, len(data)+start_columns_row):
        # 转列， A，B，C 。。。。。。
        word = string.ascii_lowercase[set_columns-1].upper()
        # 重构坐标，取数据修改
        sheet[f'{word}{str(i)}'] = data[nums]
        nums += 1

    work.save(file_path)

# 指定修改位置

if __name__ == '__main__':
    # set_excel_row('demo_xlsx.xlsx', 'Sheet1', 2, [64329, 'huang'])
    # set_excel_columns('demo_xlsx.xlsx', 'Sheet1', 2, ['张三', '李四'])
    pass


